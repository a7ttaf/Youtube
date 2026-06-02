import json
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from ums_smart_revenue.finance.bank_reconciliation import (
    MonthBankReconciliationSummary,
)
from ums_smart_revenue.finance.net_revenue import (
    ChannelNetRevenueSummary,
    MonthNetRevenueSummary,
)
from ums_smart_revenue.finance.payment_matching import MonthlyPaymentMatchSummary
from ums_smart_revenue.finance.smart_alerts import MonthlySmartAlertSummary
from ums_smart_revenue.reports.exports import ExportJobEntry
from ums_smart_revenue.reports.finance_workbook import (
    FINANCE_WORKBOOK_SHEET_NAMES,
    FinanceWorkbookPreviewValidationError,
    build_finance_workbook_preview,
    build_finance_workbook_xlsx,
)


def test_finance_workbook_preview_builds_sheet_manifest_from_source_summaries():
    """Workbook preview builds the correct sheet manifest and executive summary."""
    preview = build_finance_workbook_preview(
        export_job=_export_job(export_type="FINANCE_EXCEL"),
        net_revenue=_net_revenue_summary(),
        payment_match=_payment_match_summary(status="PAYMENT_MATCHED"),
        bank_reconciliation=_bank_summary(status="BANK_CONFIRMED"),
        smart_alerts=_smart_alert_summary(alert_count=0),
    )

    payload = preview.to_api()

    assert payload["artifact_type"] == "FINANCE_EXCEL_WORKBOOK_PREVIEW"
    assert payload["status"] == "READY_FOR_GENERATION"
    assert [sheet["name"] for sheet in payload["sheets"]] == list(
        FINANCE_WORKBOOK_SHEET_NAMES
    )
    # Provenance: every sheet that surfaces the channel-direct / account-allocated
    # deduction split must disclose the account-allocation inputs, not just the
    # dedicated Deductions sheet.
    sheet_sources = {sheet["name"]: sheet["source"] for sheet in payload["sheets"]}
    assert sheet_sources["Deductions"] == (
        "source_net_revenue_manual_overrides_deduction_components_and_account_allocations"
    )
    assert sheet_sources["Channel Breakdown"] == (
        "monthly_revenue_facts_deduction_components_and_account_allocations"
    )
    assert sheet_sources["Company Breakdown"] == (
        "monthly_revenue_facts_deduction_components_and_account_allocations"
    )
    assert sheet_sources["Sector Breakdown"] == (
        "channel_registry_revenue_facts_deduction_components_and_account_allocations"
    )
    assert payload["executive_summary"] == {
        "month": "2026-03",
        "scope_type": "company",
        "scope_id": "company-a",
        "currency": "USD",
        "month_lock_status": "LOCKED",
        "net_revenue_status": "CALCULATED",
        "payment_match_status": "PAYMENT_MATCHED",
        "bank_reconciliation_status": "BANK_CONFIRMED",
        "smart_alert_status": "CLEAR",
        "smart_alert_count": 0,
        "total_adjusted_gross_revenue_usd": "1050",
        "total_net_revenue_usd": "930",
        "total_deduction_amount_usd": "120",
        "total_channel_direct_deduction_amount_usd": "0",
        "total_account_allocated_deduction_amount_usd": "0",
        "payment_gap_usd": "0",
        "bank_gap_usd": "0",
        "channel_count": 1,
        "calculated_channel_count": 1,
    }


def test_finance_workbook_preview_rejects_non_workbook_export_type():
    """Workbook preview rejects export jobs that are not FINANCE_EXCEL."""
    with pytest.raises(FinanceWorkbookPreviewValidationError) as exc_info:
        build_finance_workbook_preview(
            export_job=_export_job(export_type="EXECUTIVE_PDF"),
            net_revenue=_net_revenue_summary(),
            payment_match=_payment_match_summary(status="PAYMENT_MATCHED"),
            bank_reconciliation=_bank_summary(status="BANK_CONFIRMED"),
            smart_alerts=_smart_alert_summary(alert_count=0),
        )

    assert str(exc_info.value) == (
        "finance workbook preview only supports FINANCE_EXCEL exports"
    )


def test_finance_workbook_xlsx_contains_expected_sheets_and_source_values():
    """Generated XLSX contains expected sheets and source values."""
    preview = build_finance_workbook_preview(
        export_job=_export_job(export_type="FINANCE_EXCEL"),
        net_revenue=_net_revenue_summary(),
        payment_match=_payment_match_summary(status="PAYMENT_MATCHED"),
        bank_reconciliation=_bank_summary(status="BANK_CONFIRMED"),
        smart_alerts=_smart_alert_summary(alert_count=0),
    )

    workbook_bytes = build_finance_workbook_xlsx(preview)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assert workbook.sheetnames == list(FINANCE_WORKBOOK_SHEET_NAMES)
    executive_summary = workbook["Executive Summary"]
    assert executive_summary["A1"].value == "Metric"
    assert executive_summary["B1"].value == "Value"
    assert executive_summary["A2"].value == "month"
    assert executive_summary["B2"].value == "2026-03"
    channel_breakdown = workbook["Channel Breakdown"]
    assert channel_breakdown["A2"].value == "channel-tv-a"
    assert channel_breakdown["H2"].value == "930"
    # The split columns exist, and a source-net channel that carries None for
    # both split fields must render as blank cells (not "0", not a crash) —
    # locks in the _decimal_to_api(None) -> blank coercion for the breakdown.
    assert channel_breakdown["J1"].value == "channel_direct_deduction_amount_usd"
    assert channel_breakdown["K1"].value == "account_allocated_deduction_amount_usd"
    assert channel_breakdown["J2"].value is None
    assert channel_breakdown["K2"].value is None
    payment_gap = workbook["Payment Gap"]
    assert payment_gap["A2"].value == "payment_match_status"
    assert payment_gap["B2"].value == "PAYMENT_MATCHED"


def _export_job(*, export_type: str) -> ExportJobEntry:
    """Build a minimal ExportJobEntry for workbook preview tests."""
    return ExportJobEntry(
        id=str(uuid4()),
        export_type=export_type,
        scope_type="company",
        scope_id="company-a",
        month="2026-03",
        currency="USD",
        requested_by=str(uuid4()),
        status="QUEUED",
        file_url=None,
        month_lock_status="LOCKED",
        include_confidence_notes=True,
        include_manual_override_notes=True,
        created_at=datetime(2026, 4, 1, tzinfo=UTC),
        completed_at=None,
    )


def _net_revenue_summary() -> MonthNetRevenueSummary:
    """Build a MonthNetRevenueSummary for workbook preview tests."""
    channel = ChannelNetRevenueSummary(
        month="2026-03",
        youtube_channel_id="channel-tv-a",
        status="CALCULATED",
        primary_source_kind="YOUTUBE_CMS",
        baseline_gross_revenue_usd=Decimal("1000.00"),
        baseline_net_revenue_usd=Decimal("880.00"),
        approved_manual_override_total_usd=Decimal("50.00"),
        adjusted_gross_revenue_usd=Decimal("1050.00"),
        net_revenue_usd=Decimal("930.00"),
        deduction_amount_usd=Decimal("120.00"),
        channel_direct_deduction_amount_usd=None,
        account_allocated_deduction_amount_usd=None,
        deduction_percentage=Decimal("11.4286"),
        confidence="B_RECONCILED",
        approved_manual_override_count=1,
        pending_manual_override_count=0,
        issues=[],
    )
    return MonthNetRevenueSummary(
        month="2026-03",
        status="CALCULATED",
        channel_count=1,
        calculated_channel_count=1,
        missing_net_source_count=0,
        pending_manual_override_count=0,
        total_adjusted_gross_revenue_usd=Decimal("1050.00"),
        total_net_revenue_usd=Decimal("930.00"),
        total_deduction_amount_usd=Decimal("120.00"),
        total_channel_direct_deduction_amount_usd=Decimal("0.00"),
        total_account_allocated_deduction_amount_usd=Decimal("0.00"),
        unallocated_account_deduction_total_usd=None,
        unallocated_account_issues=None,
        channels=[channel],
    )


def _payment_match_summary(*, status: str) -> MonthlyPaymentMatchSummary:
    """Build a MonthlyPaymentMatchSummary for workbook preview tests."""
    return MonthlyPaymentMatchSummary(
        month="2026-03",
        currency="USD",
        status=status,
        youtube_revenue_total_usd=Decimal("930.00"),
        adsense_paid_amount=Decimal("930.00"),
        payment_gap_usd=Decimal("0.00"),
        youtube_source_channel_count=1,
        missing_youtube_source_channel_count=0,
        payment_count=1,
        paid_payment_count=1,
        non_paid_payment_count=0,
        unsupported_payment_currency_count=0,
        tolerance_usd=Decimal("0.01"),
        issues=[],
    )


def _bank_summary(*, status: str) -> MonthBankReconciliationSummary:
    """Build a MonthBankReconciliationSummary for workbook preview tests."""
    return MonthBankReconciliationSummary(
        month="2026-03",
        currency="USD",
        status=status,
        adsense_paid_amount_usd=Decimal("930.00"),
        bank_received_amount_usd=Decimal("930.00"),
        bank_gap_usd=Decimal("0.00"),
        transfer_fee_usd=Decimal("0.00"),
        fx_difference_usd=Decimal("0.00"),
        payment_count=1,
        paid_payment_count=1,
        non_paid_payment_count=0,
        unsupported_payment_currency_count=0,
        entry_count=1,
        tolerance_usd=Decimal("0.01"),
        issues=[],
        entries=[],
    )


def _smart_alert_summary(*, alert_count: int) -> MonthlySmartAlertSummary:
    """Build a MonthlySmartAlertSummary for workbook preview tests."""
    assert alert_count == 0
    return MonthlySmartAlertSummary(
        month="2026-03",
        status="CLEAR",
        highest_severity=None,
        alerts=[],
    )


def _net_revenue_summary_with_breakdown() -> MonthNetRevenueSummary:
    """Build a summary with a COMPONENT_DERIVED channel carrying a real split."""
    channel = ChannelNetRevenueSummary(
        month="2026-03",
        youtube_channel_id="channel-tv-a",
        status="COMPONENT_DERIVED",
        primary_source_kind="ADSENSE",
        baseline_gross_revenue_usd=Decimal("1000.00"),
        baseline_net_revenue_usd=None,
        approved_manual_override_total_usd=Decimal("0.00"),
        adjusted_gross_revenue_usd=Decimal("1000.00"),
        net_revenue_usd=Decimal("870.00"),
        deduction_amount_usd=Decimal("130.00"),
        channel_direct_deduction_amount_usd=Decimal("30.00"),
        account_allocated_deduction_amount_usd=Decimal("100.00"),
        deduction_percentage=Decimal("13.0000"),
        confidence="D_ESTIMATED",
        approved_manual_override_count=0,
        pending_manual_override_count=0,
        issues=[],
    )
    return MonthNetRevenueSummary(
        month="2026-03",
        status="CALCULATED",
        channel_count=1,
        calculated_channel_count=1,
        missing_net_source_count=0,
        pending_manual_override_count=0,
        total_adjusted_gross_revenue_usd=Decimal("1000.00"),
        total_net_revenue_usd=Decimal("870.00"),
        total_deduction_amount_usd=Decimal("130.00"),
        total_channel_direct_deduction_amount_usd=Decimal("30.00"),
        total_account_allocated_deduction_amount_usd=Decimal("100.00"),
        unallocated_account_deduction_total_usd=None,
        unallocated_account_issues=None,
        channels=[channel],
    )


def test_finance_workbook_renders_deduction_breakdown_columns_and_rows():
    """XLSX renders per-channel split columns, aggregate summary/scope rows, and Raw Appendix."""
    preview = build_finance_workbook_preview(
        export_job=_export_job(export_type="FINANCE_EXCEL"),
        net_revenue=_net_revenue_summary_with_breakdown(),
        payment_match=_payment_match_summary(status="PAYMENT_MATCHED"),
        bank_reconciliation=_bank_summary(status="BANK_CONFIRMED"),
        smart_alerts=_smart_alert_summary(alert_count=0),
    )
    workbook = load_workbook(BytesIO(build_finance_workbook_xlsx(preview)), data_only=True)

    channel_breakdown = workbook["Channel Breakdown"]
    assert channel_breakdown["H1"].value == "net_revenue_usd"
    assert channel_breakdown["I1"].value == "deduction_amount_usd"
    assert channel_breakdown["J1"].value == "channel_direct_deduction_amount_usd"
    assert channel_breakdown["K1"].value == "account_allocated_deduction_amount_usd"
    assert channel_breakdown["I2"].value == "130"
    assert channel_breakdown["J2"].value == "30"
    assert channel_breakdown["K2"].value == "100"

    deductions = workbook["Deductions"]
    assert deductions["D1"].value == "deduction_amount_usd"
    assert deductions["E1"].value == "channel_direct_deduction_amount_usd"
    assert deductions["F1"].value == "account_allocated_deduction_amount_usd"
    assert deductions["D2"].value == "130"
    assert deductions["E2"].value == "30"
    assert deductions["F2"].value == "100"

    exec_summary = preview.to_api()["executive_summary"]
    assert exec_summary["total_channel_direct_deduction_amount_usd"] == "30"
    assert exec_summary["total_account_allocated_deduction_amount_usd"] == "100"

    for sheet_name in ("Company Breakdown", "Sector Breakdown"):
        keys = {row[0].value: row[1].value for row in workbook[sheet_name].iter_rows()}
        assert keys["total_channel_direct_deduction_amount_usd"] == "30"
        assert keys["total_account_allocated_deduction_amount_usd"] == "100"

    raw = {row[0].value: row[1].value for row in workbook["Raw Appendix"].iter_rows()}
    net_revenue_json = json.loads(raw["net_revenue"])
    assert net_revenue_json["total_channel_direct_deduction_amount_usd"] == "30"
    assert net_revenue_json["total_account_allocated_deduction_amount_usd"] == "100"
