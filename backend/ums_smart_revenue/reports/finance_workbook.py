import json
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from ums_smart_revenue.finance.bank_reconciliation import (
    MonthBankReconciliationSummary,
)
from ums_smart_revenue.finance.decimal_formatting import decimal_to_api as _decimal_to_api
from ums_smart_revenue.finance.net_revenue import MonthNetRevenueSummary
from ums_smart_revenue.finance.payment_matching import MonthlyPaymentMatchSummary
from ums_smart_revenue.finance.smart_alerts import MonthlySmartAlertSummary
from ums_smart_revenue.reports.exports import ExportJobEntry

FINANCE_WORKBOOK_SHEET_NAMES = (
    "Executive Summary",
    "Monthly Close",
    "Company Breakdown",
    "Sector Breakdown",
    "Channel Breakdown",
    "Deductions",
    "Payment Gap",
    "Confidence Notes",
    "Raw Appendix",
)

_SHEET_SOURCES = {
    "Executive Summary": "computed_finance_summary",
    "Monthly Close": "finance_month_close",
    "Company Breakdown": "monthly_revenue_facts_deduction_components_and_account_allocations",
    "Sector Breakdown": (
        "channel_registry_revenue_facts_deduction_components_and_account_allocations"
    ),
    "Channel Breakdown": "monthly_revenue_facts_deduction_components_and_account_allocations",
    "Deductions": (
        "source_net_revenue_manual_overrides_deduction_components_and_account_allocations"
    ),
    "Payment Gap": "adsense_payments_and_bank_reconciliation",
    "Confidence Notes": "confidence_and_smart_alerts",
    "Raw Appendix": "source_record_references",
}


class FinanceWorkbookPreviewValidationError(ValueError):
    """Exception raised when validation of a finance workbook preview fails."""


@dataclass(frozen=True)
class FinanceWorkbookSheet:
    """Data class representing a single sheet in the finance workbook preview."""

    name: str
    source: str
    status: str
    sensitive: bool

    def to_api(self) -> dict[str, object]:
        """Convert the object's attributes to a dictionary suitable for API use.

        Returns:
            dict[str, object]: A dictionary containing the "name", "source", "status",
                and "sensitive" attributes of this object.
        """
        return {
            "name": self.name,
            "source": self.source,
            "status": self.status,
            "sensitive": self.sensitive,
        }


@dataclass(frozen=True)
class FinanceWorkbookPreview:
    """Finance workbook preview with export metadata and source summaries."""

    export_job: ExportJobEntry
    sheets: tuple[FinanceWorkbookSheet, ...]
    net_revenue: MonthNetRevenueSummary
    payment_match: MonthlyPaymentMatchSummary
    bank_reconciliation: MonthBankReconciliationSummary
    smart_alerts: MonthlySmartAlertSummary

    def to_api(self) -> dict[str, object]:
        """Convert this FinanceWorkbookPreview into a dictionary payload for API export."""
        return {
            "export_id": self.export_job.id,
            "export_type": self.export_job.export_type,
            "artifact_type": "FINANCE_EXCEL_WORKBOOK_PREVIEW",
            "status": "READY_FOR_GENERATION",
            "month": self.export_job.month,
            "currency": self.export_job.currency,
            "scope_type": self.export_job.scope_type,
            "scope_id": self.export_job.scope_id,
            "month_lock_status": self.export_job.month_lock_status,
            "include_confidence_notes": self.export_job.include_confidence_notes,
            "include_manual_override_notes": (
                self.export_job.include_manual_override_notes
            ),
            "sheets": [sheet.to_api() for sheet in self.sheets],
            "executive_summary": _executive_summary(
                export_job=self.export_job,
                net_revenue=self.net_revenue,
                payment_match=self.payment_match,
                bank_reconciliation=self.bank_reconciliation,
                smart_alerts=self.smart_alerts,
            ),
            "source_summaries": {
                "net_revenue": self.net_revenue.to_api(),
                "payment_match": self.payment_match.to_api(),
                "bank_reconciliation": self.bank_reconciliation.to_api(),
                "smart_alerts": self.smart_alerts.to_api(),
            },
        }


def build_finance_workbook_preview(
    *,
    export_job: ExportJobEntry,
    net_revenue: MonthNetRevenueSummary,
    payment_match: MonthlyPaymentMatchSummary,
    bank_reconciliation: MonthBankReconciliationSummary,
    smart_alerts: MonthlySmartAlertSummary,
) -> FinanceWorkbookPreview:
    """Build a finance workbook preview from export metadata and summaries."""
    if export_job.export_type != "FINANCE_EXCEL":
        raise FinanceWorkbookPreviewValidationError(
            "finance workbook preview only supports FINANCE_EXCEL exports"
        )
    _validate_same_month(
        export_job=export_job,
        net_revenue=net_revenue,
        payment_match=payment_match,
        bank_reconciliation=bank_reconciliation,
        smart_alerts=smart_alerts,
    )
    if export_job.currency != "USD":
        raise FinanceWorkbookPreviewValidationError(
            "finance workbook preview requires USD export currency"
        )
    if payment_match.currency != "USD" or bank_reconciliation.currency != "USD":
        raise FinanceWorkbookPreviewValidationError(
            "finance workbook preview requires USD source summaries"
        )

    return FinanceWorkbookPreview(
        export_job=export_job,
        sheets=tuple(
            FinanceWorkbookSheet(
                name=name,
                source=_SHEET_SOURCES[name],
                status="READY",
                sensitive=True,
            )
            for name in FINANCE_WORKBOOK_SHEET_NAMES
        ),
        net_revenue=net_revenue,
        payment_match=payment_match,
        bank_reconciliation=bank_reconciliation,
        smart_alerts=smart_alerts,
    )


def build_finance_workbook_xlsx(preview: FinanceWorkbookPreview) -> bytes:
    """Generate an XLSX workbook binary from a finance workbook preview."""
    workbook = Workbook()
    workbook.iso_dates = True
    summary_sheet = workbook.active
    summary_sheet.title = FINANCE_WORKBOOK_SHEET_NAMES[0]

    _write_key_value_sheet(
        summary_sheet,
        _executive_summary(
            export_job=preview.export_job,
            net_revenue=preview.net_revenue,
            payment_match=preview.payment_match,
            bank_reconciliation=preview.bank_reconciliation,
            smart_alerts=preview.smart_alerts,
        ),
    )
    _write_key_value_sheet(
        workbook.create_sheet("Monthly Close"),
        {
            "month": preview.export_job.month,
            "month_lock_status": preview.export_job.month_lock_status,
            "export_job_status": preview.export_job.status,
            "include_confidence_notes": preview.export_job.include_confidence_notes,
            "include_manual_override_notes": (
                preview.export_job.include_manual_override_notes
            ),
        },
    )
    _write_key_value_sheet(
        workbook.create_sheet("Company Breakdown"),
        _scope_breakdown(preview, breakdown_type="company"),
    )
    _write_key_value_sheet(
        workbook.create_sheet("Sector Breakdown"),
        _scope_breakdown(preview, breakdown_type="sector"),
    )
    _write_table_sheet(
        workbook.create_sheet("Channel Breakdown"),
        [
            "youtube_channel_id",
            "status",
            "primary_source_kind",
            "baseline_gross_revenue_usd",
            "baseline_net_revenue_usd",
            "approved_manual_override_total_usd",
            "adjusted_gross_revenue_usd",
            "net_revenue_usd",
            "deduction_amount_usd",
            "channel_direct_deduction_amount_usd",
            "account_allocated_deduction_amount_usd",
            "deduction_percentage",
            "confidence",
            "approved_manual_override_count",
            "pending_manual_override_count",
        ],
        [
            [
                channel.youtube_channel_id,
                channel.status,
                channel.primary_source_kind,
                _decimal_to_api(channel.baseline_gross_revenue_usd),
                _decimal_to_api(channel.baseline_net_revenue_usd),
                _decimal_to_api(channel.approved_manual_override_total_usd),
                _decimal_to_api(channel.adjusted_gross_revenue_usd),
                _decimal_to_api(channel.net_revenue_usd),
                _decimal_to_api(channel.deduction_amount_usd),
                _decimal_to_api(channel.channel_direct_deduction_amount_usd),
                _decimal_to_api(channel.account_allocated_deduction_amount_usd),
                _decimal_to_api(channel.deduction_percentage),
                channel.confidence,
                channel.approved_manual_override_count,
                channel.pending_manual_override_count,
            ]
            for channel in preview.net_revenue.channels
        ],
    )
    _write_table_sheet(
        workbook.create_sheet("Deductions"),
        [
            "youtube_channel_id",
            "adjusted_gross_revenue_usd",
            "net_revenue_usd",
            "deduction_amount_usd",
            "channel_direct_deduction_amount_usd",
            "account_allocated_deduction_amount_usd",
            "deduction_percentage",
            "approved_manual_override_total_usd",
        ],
        [
            [
                channel.youtube_channel_id,
                _decimal_to_api(channel.adjusted_gross_revenue_usd),
                _decimal_to_api(channel.net_revenue_usd),
                _decimal_to_api(channel.deduction_amount_usd),
                _decimal_to_api(channel.channel_direct_deduction_amount_usd),
                _decimal_to_api(channel.account_allocated_deduction_amount_usd),
                _decimal_to_api(channel.deduction_percentage),
                _decimal_to_api(channel.approved_manual_override_total_usd),
            ]
            for channel in preview.net_revenue.channels
        ],
    )
    _write_key_value_sheet(
        workbook.create_sheet("Payment Gap"),
        {
            "payment_match_status": preview.payment_match.status,
            "youtube_revenue_total_usd": _decimal_to_api(
                preview.payment_match.youtube_revenue_total_usd
            ),
            "adsense_paid_amount": _decimal_to_api(
                preview.payment_match.adsense_paid_amount
            ),
            "payment_gap_usd": _decimal_to_api(preview.payment_match.payment_gap_usd),
            "bank_reconciliation_status": preview.bank_reconciliation.status,
            "bank_received_amount_usd": _decimal_to_api(
                preview.bank_reconciliation.bank_received_amount_usd
            ),
            "bank_gap_usd": _decimal_to_api(preview.bank_reconciliation.bank_gap_usd),
            "transfer_fee_usd": _decimal_to_api(
                preview.bank_reconciliation.transfer_fee_usd
            ),
            "fx_difference_usd": _decimal_to_api(
                preview.bank_reconciliation.fx_difference_usd
            ),
        },
    )
    _write_table_sheet(
        workbook.create_sheet("Confidence Notes"),
        ["code", "severity", "source", "confidence", "message"],
        [
            [
                alert.code,
                alert.severity,
                alert.source,
                alert.confidence,
                alert.message,
            ]
            for alert in preview.smart_alerts.alerts
        ]
        or [["CLEAR", None, "smart_alerts", None, "No smart alerts for this scope."]],
    )
    _write_key_value_sheet(
        workbook.create_sheet("Raw Appendix"),
        {
            source_name: json.dumps(payload, sort_keys=True)
            for source_name, payload in preview.to_api()["source_summaries"].items()
        },
    )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _validate_same_month(
    *,
    export_job: ExportJobEntry,
    net_revenue: MonthNetRevenueSummary,
    payment_match: MonthlyPaymentMatchSummary,
    bank_reconciliation: MonthBankReconciliationSummary,
    smart_alerts: MonthlySmartAlertSummary,
) -> None:
    """
    Ensure that all provided summaries correspond to the same month as the export job.
    Raises a validation error if multiple different months are found.
    """
    months = {
        export_job.month,
        net_revenue.month,
        payment_match.month,
        bank_reconciliation.month,
        smart_alerts.month,
    }
    if len(months) != 1:
        raise FinanceWorkbookPreviewValidationError(
            "finance workbook preview source summaries must use the export month"
        )


def _executive_summary(
    *,
    export_job: ExportJobEntry,
    net_revenue: MonthNetRevenueSummary,
    payment_match: MonthlyPaymentMatchSummary,
    bank_reconciliation: MonthBankReconciliationSummary,
    smart_alerts: MonthlySmartAlertSummary,
) -> dict[str, object]:
    """
    Build and return the executive summary dictionary combining export job info
    and statuses and metrics from net revenue, payment match, bank reconciliation,
    and smart alerts summaries.
    """
    return {
        "month": export_job.month,
        "scope_type": export_job.scope_type,
        "scope_id": export_job.scope_id,
        "currency": export_job.currency,
        "month_lock_status": export_job.month_lock_status,
        "net_revenue_status": net_revenue.status,
        "payment_match_status": payment_match.status,
        "bank_reconciliation_status": bank_reconciliation.status,
        "smart_alert_status": smart_alerts.status,
        "smart_alert_count": len(smart_alerts.alerts),
        "total_adjusted_gross_revenue_usd": _decimal_to_api(
            net_revenue.total_adjusted_gross_revenue_usd
        ),
        "total_net_revenue_usd": _decimal_to_api(net_revenue.total_net_revenue_usd),
        "total_deduction_amount_usd": _decimal_to_api(
            net_revenue.total_deduction_amount_usd
        ),
        "total_channel_direct_deduction_amount_usd": _decimal_to_api(
            net_revenue.total_channel_direct_deduction_amount_usd
        ),
        "total_account_allocated_deduction_amount_usd": _decimal_to_api(
            net_revenue.total_account_allocated_deduction_amount_usd
        ),
        "payment_gap_usd": _decimal_to_api(payment_match.payment_gap_usd),
        "bank_gap_usd": _decimal_to_api(bank_reconciliation.bank_gap_usd),
        "channel_count": net_revenue.channel_count,
        "calculated_channel_count": net_revenue.calculated_channel_count,
    }


def _scope_breakdown(
    preview: FinanceWorkbookPreview, *, breakdown_type: str
) -> dict[str, object]:
    """
    Generate a scope breakdown dictionary for the given preview and breakdown type,
    including revenue metrics and channel counts.
    """
    return {
        "breakdown_type": breakdown_type,
        "scope_type": preview.export_job.scope_type,
        "scope_id": preview.export_job.scope_id,
        "month": preview.export_job.month,
        "currency": preview.export_job.currency,
        "channel_count": preview.net_revenue.channel_count,
        "calculated_channel_count": preview.net_revenue.calculated_channel_count,
        "total_adjusted_gross_revenue_usd": _decimal_to_api(
            preview.net_revenue.total_adjusted_gross_revenue_usd
        ),
        "total_net_revenue_usd": _decimal_to_api(
            preview.net_revenue.total_net_revenue_usd
        ),
        "total_deduction_amount_usd": _decimal_to_api(
            preview.net_revenue.total_deduction_amount_usd
        ),
        "total_channel_direct_deduction_amount_usd": _decimal_to_api(
            preview.net_revenue.total_channel_direct_deduction_amount_usd
        ),
        "total_account_allocated_deduction_amount_usd": _decimal_to_api(
            preview.net_revenue.total_account_allocated_deduction_amount_usd
        ),
    }


def _write_key_value_sheet(sheet: Worksheet, values: dict[str, object]) -> None:
    """
    Write a key-value mapping to the provided worksheet as a two-column table.
    Keys become the first column and their corresponding formatted values the second.
    """
    _write_table_sheet(
        sheet,
        ["Metric", "Value"],
        [[key, _cell_value(value)] for key, value in values.items()],
    )


def _write_table_sheet(
    sheet: Worksheet, headers: list[str], rows: list[list[object]]
) -> None:
    """
    Populate the worksheet with the given headers and rows,
    then freeze the header row, apply header styling, and autosize columns.
    """
    sheet.append(headers)
    for row in rows:
        sheet.append([_cell_value(value) for value in row])
    sheet.freeze_panes = "A2"
    _style_header(sheet)
    _autosize_columns(sheet)


def _style_header(sheet: Worksheet) -> None:
    """
    Apply styling to the header row of the sheet: solid fill, white bold font,
    and center alignment.
    """
    fill = PatternFill("solid", fgColor="1F2933")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(sheet: Worksheet) -> None:
    """
    Adjust each column's width based on the maximum length of cell values,
    with defined minimum and maximum width constraints.
    """
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def _cell_value(value: object) -> object:
    """
    Convert values for cell insertion: Decimal to API format,
    dict or list to JSON string, otherwise return value unchanged.
    """
    if isinstance(value, Decimal):
        return _decimal_to_api(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True)
    return value
